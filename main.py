import dwdpollen, datetime, openpyxl, os, json, time, schedule
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle

api = dwdpollen.DwdPollenApi()

def load_config():
    with open('config.json', 'r') as f:
        config = json.load(f)
    return config

class pollenType:
    def __init__(self, name, severity):
        self.name = name
        self.severity = severity

    def __repr__(self):
        return f"pollenType(name={self.name}, severity={self.severity})"
    
config = load_config()

today = datetime.date.today().isoformat()
filename='Heuschnupfentagebuch.xlsx'

def getPollenList():
    region = api.get_pollen(config.get("region_id"), config.get("partregion_id"))
    dailyPollen = region.get('pollen')

    def getPollenType(name):
        return pollenType(name, dailyPollen.get(name).get(today).get('value'))
    
    pollenList = [ getPollenType('Hasel'), 
    getPollenType('Erle'), 
    getPollenType('Birke'),
    getPollenType('Graeser'),
    getPollenType('Esche'),
    getPollenType('Ambrosia'),
    getPollenType('Roggen'),
    getPollenType('Beifuss')]
    return pollenList
print("ENDE")



def schreibe_in_excel(pollenList, ws, wb):

    stratRow = ws.max_row + 1
        
    ws.cell(row=stratRow, column=1, value=today).number_format = 'yyyy-mm-dd;@'
    ws.cell(row=stratRow, column=2, value="")

    for col_num, pollentyp in enumerate(pollenList, start=4):
        ws.cell(row=stratRow, column=col_num, value=pollentyp.severity).alignment = Alignment(horizontal='center', vertical='center')
        
    wb.save(filename)

def run_script():
    if os.path.exists(filename):

        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        writeAllowed = True
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == today:
                print("Entry for today exists. Skipping execution")
                writeAllowed = False
        if writeAllowed == True:
            pollenList = getPollenList()
            schreibe_in_excel(pollenList, ws, wb)
    else:
        wb = Workbook()
        ws = wb.active
        pollenList = getPollenList()
        colnames = ['Datum', 'Heuschnupfengrad', 'Medizin genommen'] + [pollentyp.name for pollentyp in pollenList]
        ws.append(colnames)
        schreibe_in_excel(pollenList, ws, wb)

def main():
    # initial run on fiirst start
    run_script()
    
    interval_hours = config.get("interval_hours", 1)

    schedule.every(interval_hours).hours.do(run_script)

    print(f"Scheduled job to run every {interval_hours} hours.")

    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == "__main__":
    main()
